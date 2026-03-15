from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.config import Settings
from app.db.session import get_mongo_database
from app.infrastructure.database_gateway import DatabaseGateway


class ExcelExportService:
    def __init__(self, settings: Settings) -> None:
        self._export_path = settings.excel_export_path

    def export_user_workspace(self, user_id: str) -> str:
        Path(self._export_path).parent.mkdir(parents=True, exist_ok=True)
        gateway = DatabaseGateway(get_mongo_database())
        charts = gateway.list_saved_charts_for_user(user_id, limit=5000, status="all")
        comparisons = gateway.list_saved_comparisons_for_user(user_id, limit=5000, status="all")

        wb = Workbook()
        ws_saved_charts = wb.active
        ws_saved_charts.title = "Saved Charts"
        self._write_rows(
            ws_saved_charts,
            [
                "name",
                "city",
                "birth_date",
                "birth_time",
                "timezone_offset_minutes",
                "country",
                "state",
                "town",
                "latitude",
                "longitude",
                "time_zone_id",
                "ascendant_sign",
                "notes",
                "saved_at",
                "updated_at",
                "archived_at",
            ],
            [chart.model_dump(mode="json") for chart in charts],
        )

        ws_saved_comparisons = wb.create_sheet("Saved Comparisons")
        self._write_rows(
            ws_saved_comparisons,
            [
                "primary_name",
                "partner_name",
                "compatibility_score",
                "summary",
                "notes",
                "saved_at",
                "updated_at",
                "archived_at",
            ],
            [comparison.model_dump(mode="json") for comparison in comparisons],
        )

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B3144", end_color="1B3144", fill_type="solid")

        for ws in wb.worksheets:
            if ws.max_row and ws[1]:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    max_length = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

        wb.save(self._export_path)
        return self._export_path

    @staticmethod
    def _write_rows(ws, columns: list[str], rows: list[dict]) -> None:
        ws.append(columns)
        for row in rows:
            ws.append([row.get(column) for column in columns])
